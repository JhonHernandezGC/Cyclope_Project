"""
cyclope_core.py
===============
Lógica del pipeline Cyclope como funciones reutilizables (sin rutas fijas):

  run_extraction(pdf_paths, output_dir, log)            -> genera camara_extraction.xlsx/.json
  run_prefill(excel_path, templates_dir, output_dir, log) -> genera los 7 documentos por empresa
  create_counterparty_folders(excel_path, network_base, output_dir, log) -> carpetas en la red

`log` es una función callback (por defecto print) para enviar el progreso a la GUI.

Las funciones de extracción y de llenado provienen verbatim de extract_camara.py
y pre_fill.ipynb.
"""
from __future__ import annotations

import json
import re
import shutil
import unicodedata
import zipfile
from dataclasses import dataclass, asdict
from datetime import date
from pathlib import Path
from typing import Any, Callable

import pandas as pd
import pdfplumber
from lxml import etree
from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter

BOGOTA = "Bogotá D.C."
W_NS   = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
XML_NS = "http://www.w3.org/XML/1998/namespace"

# Nombres de archivo de las plantillas (dentro de la carpeta Requirements)
TEMPLATE_FILES = {
    "annex":   "Annex A - Signature Card for Accounts in COP - Colombia (English and Spanish).docx",
    "contact": "Contact Information Certificate For Electronic Signatures.docx",
    "crs":     "CRS FATCA Form.pdf",
    "esst":    "Electronic Signatures Services Terms.docx",
    "ge":      "Formulario Grandes Exposiciones.xlsx",
    "gcc":     "GCC-Declaracion de Veracidad.pdf",
    "kyc":     "KYC application form (Onboarding) v. Mar 2026.pdf",
}


def _safe_folder_name(name: str) -> str:
    """Elimina caracteres no permitidos en nombres de carpeta en Windows."""
    return re.sub(r'[/:*?"<>|]', "_", name).strip()


# ===========================================================================
# SECCIÓN 1 — Extracción (verbatim de extract_camara.py)
# ===========================================================================
DOC_ID_RE = re.compile(r"\b(C\.?\s*C\.?|C\.?\s*E\.?|P\.?\s*P\.?)\s*No\.?\s*([A-Z0-9.\-]+)", re.I)

STOP_HEADINGS = {
    "CAMARA DE COMERCIO DE BOGOTA",
    "SEDE VIRTUAL",
    "CERTIFICADO DE EXISTENCIA Y REPRESENTACION LEGAL",
    "CERTIFICADO DE INSCRIPCION DE DOCUMENTOS",
    "FECHA EXPEDICION",
    "RECIBO NO",
    "VALOR",
    "CODIGO DE VERIFICACION",
    "VERIFIQUE EL CONTENIDO",
    "PAGINA",
    "NOMBRAMIENTOS",
    "ORGANO DE ADMINISTRACION",
    "JUNTA DIRECTIVA",
    "REPRESENTANTES LEGALES",
    "REFORMAS DE ESTATUTOS",
    "REVISORES FISCALES",
    "PODERES",
    "DOCUMENTO INSCRIPCION",
    "CAPITAL",
    "MATRICULA",
    "UBICACION",
}


@dataclass
class Appointment:
    section: str
    group: str
    cargo: str
    nombre: str
    identificacion_tipo: str
    identificacion_numero: str
    page: int


def strip_accents(value: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", value) if not unicodedata.combining(c)
    )


def norm(value: str) -> str:
    value = strip_accents(value).upper()
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_space(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def extract_pages(pdf_path: Path) -> list[dict[str, Any]]:
    pages: list[dict[str, Any]] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False)
            pages.append({"page": index, "text": text, "words": words})
    return pages


def extract_basic_info(full_text: str) -> dict[str, str]:
    patterns = {
        "razon_social": r"Raz[oó]n social:\s*(.+)",
        "nit": r"\bNit:\s*([0-9.\-\s]+)",
        "domicilio_principal": r"Domicilio principal:\s*(.+)",
        "direccion_domicilio_principal": r"Direcci[oó]n\s+del\s+domicilio\s+principal\s*:\s*(.*?)(?:\n|$)", #se cruza con linea 337
        "Teléfono comercial 1:": r"Tel[eé]fono comercial 1:\s*(.+)", # se cruza con linea 337
        "ciiu": r"C[oó]digo\s+CIIU\s*[:\-]?\s*(\d{4})"
    }
    result: dict[str, str] = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, full_text, flags=re.I)
        result[key] = clean_space(match.group(1)) if match else ""
    result["nit"] = re.sub(r"\s+", "-", result.get("nit", "")).strip("-")
    return result


def normalized_index(original: str) -> tuple[str, list[int]]:
    normalized_chars: list[str] = []
    index_map: list[int] = []
    for original_index, char in enumerate(original):
        decomposed = unicodedata.normalize("NFKD", char)
        for part in decomposed:
            if not unicodedata.combining(part):
                normalized_chars.append(part.upper())
                index_map.append(original_index)
    return "".join(normalized_chars), index_map


def section_between(full_text: str, start_terms: list[str], end_terms: list[str]) -> str:
    normalized, index_map = normalized_index(full_text)
    starts = [(normalized.find(term), term) for term in start_terms]
    starts = [(pos, term) for pos, term in starts if pos >= 0]
    if not starts:
        return ""
    start_pos, start_term = min(starts, key=lambda item: item[0])
    start_original = index_map[start_pos + len(start_term) - 1] + 1

    end_candidates: list[int] = []
    for term in end_terms:
        pos = normalized.find(term, start_pos + len(start_term))
        if pos >= 0:
            end_candidates.append(index_map[pos])
    end_original = min(end_candidates) if end_candidates else len(full_text)
    return clean_space(full_text[start_original:end_original])


def extract_faculties(full_text: str) -> dict[str, Any]:
    faculties_text = section_between(
        full_text,
        ["FACULTADES Y LIMITACIONES DEL REPRESENTANTE LEGAL", "FACULTADES Y LIMITACIONES"],
        ["NOMBRAMIENTOS", "REFORMAS DE ESTATUTOS", "RECURSOS", "CERTIFICA"],
    )
    limitations_area = norm(faculties_text if faculties_text else full_text)
    judicial_area = norm(full_text)

    limitation_patterns = [
        r"AUTORIZACION EXPRESA",
        r"CUANTIA",
        r"SUPERIOR A",
        r"NECESITAR[A-Z]* AUTORIZACION",
        r"LIMITACION",
        r"NO PODRA",
        r"JUNTA DE (MIEMBROS|SOCIOS|ACCIONISTAS|DIRECTIVA)",
    ]
    judicial_patterns = [
        r"REPRESENTANTES? JUDICIALES?",
        r"APODERAD[OA]S?",
        r"PODER GENERAL",
        r"ACTUACIONES JUDICIALES",
        r"AUDIENCIAS? DE CONCILIACION",
    ]

    limitations_found = sorted(
        {pattern for pattern in limitation_patterns if re.search(pattern, limitations_area)}
    )
    judicial_found = sorted(
        {pattern for pattern in judicial_patterns if re.search(pattern, judicial_area)}
    )
    return {
        "facultades_texto": faculties_text,
        "tiene_limitaciones": bool(limitations_found),
        "reglas_limitaciones_detectadas": limitations_found,
        "tiene_indicios_representacion_judicial": bool(judicial_found),
        "reglas_representacion_judicial_detectadas": judicial_found,
    }


def group_words_by_line(words: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
    lines: dict[int, list[dict[str, Any]]] = {}
    for word in words:
        top = round(float(word["top"]) / 3) * 3
        lines.setdefault(top, []).append(word)
    return [sorted(lines[top], key=lambda item: item["x0"]) for top in sorted(lines)]


def line_text(line: list[dict[str, Any]]) -> str:
    return clean_space(" ".join(str(word["text"]) for word in line))


def line_columns(line: list[dict[str, Any]]) -> tuple[str, str, str]:
    cargo_words: list[str] = []
    name_words: list[str] = []
    id_words: list[str] = []
    for word in line:
        x0 = float(word["x0"])
        text = str(word["text"])
        if x0 < 205:
            cargo_words.append(text)
        elif x0 < 370:
            name_words.append(text)
        else:
            id_words.append(text)
    return clean_space(" ".join(cargo_words)), clean_space(" ".join(name_words)), clean_space(" ".join(id_words))


def is_stop_line(raw: str) -> bool:
    normalized = norm(raw)
    if not normalized:
        return True
    if normalized.startswith("PAGINA "):
        return True
    if normalized.startswith(("POR ACTA", "POR DOCUMENTO", "POR ESCRITURA", "INSCRITA ")):
        return True
    return any(normalized.startswith(heading) for heading in STOP_HEADINGS)


def extract_appointments(pages: list[dict[str, Any]]) -> list[Appointment]:
    appointments: list[Appointment] = []
    active_table = False
    section = ""
    group = ""
    current: Appointment | None = None

    def flush_current() -> None:
        nonlocal current
        if current:
            current.cargo = clean_space(current.cargo)
            current.nombre = clean_space(current.nombre)
            appointments.append(current)
            current = None

    for page in pages:
        page_no = int(page["page"])
        for line in group_words_by_line(page["words"]):
            raw = line_text(line)
            normalized = norm(raw)

            if normalized in {"REPRESENTANTES LEGALES", "REPRESENTANTE LEGAL"}:
                flush_current()
                section = "REPRESENTANTES LEGALES"
                group = ""
                active_table = False
                continue
            if normalized in {"ORGANO DE ADMINISTRACION", "JUNTA DIRECTIVA"}:
                flush_current()
                section = "ORGANO DE ADMINISTRACION"
                active_table = False
                continue
            if normalized in {"REVISORES FISCALES", "PODERES"}:
                flush_current()
                section = ""
                group = ""
                active_table = False
                continue
            if normalized in {"PRINCIPALES", "SUPLENTES"}:
                flush_current()
                group = normalized
                active_table = False
                continue
            if "CARGO" in normalized and "NOMBRE" in normalized and "IDENTIFICACION" in normalized:
                flush_current()
                active_table = section in {"REPRESENTANTES LEGALES", "ORGANO DE ADMINISTRACION"}
                continue

            if not active_table:
                continue

            if is_stop_line(raw):
                flush_current()
                active_table = False
                continue

            cargo, name, ident = line_columns(line)
            id_match = DOC_ID_RE.search(ident)
            if id_match:
                flush_current()
                current = Appointment(
                    section=section or "NOMBRAMIENTOS",
                    group=group,
                    cargo=cargo,
                    nombre=name,
                    identificacion_tipo=clean_space(id_match.group(1).replace(" ", "")),
                    identificacion_numero=clean_space(id_match.group(2)),
                    page=page_no,
                )
                continue

            if current and not is_stop_line(raw):
                if cargo:
                    current.cargo = clean_space(current.cargo + " " + cargo)
                if name:
                    current.nombre = clean_space(current.nombre + " " + name)

        flush_current()
        active_table = False

    flush_current()
    return appointments


def extract_pdf(pdf_path: Path) -> dict[str, Any]:
    pages = extract_pages(pdf_path)
    full_text = "\n".join(page["text"] for page in pages)
    basic = extract_basic_info(full_text)
    appointments = extract_appointments(pages)
    return {
        "archivo": str(pdf_path),
        **basic,
        **extract_faculties(full_text),
        "nombramientos": [asdict(item) for item in appointments],
    }


def write_outputs(records: list[dict[str, Any]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "camara_extraction.json").write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    summary_rows = []
    appointment_rows = []
    for record in records:
        summary_rows.append(
            {
                "archivo": Path(record["archivo"]).name,
                "razon_social": record["razon_social"],
                "nit": record["nit"],
                "domicilio_principal": record["domicilio_principal"],
                "direccion_domicilio_principal": record["direccion_domicilio_principal"],  # nueva columna localizada en linea 102
                "Teléfono comercial 1:": record["Teléfono comercial 1:"],  # nueva columna localizada en linea 103
                "Actividad principal Código CIIU": record.get("ciiu", ""),
                "tiene_limitaciones": record["tiene_limitaciones"],
                "tiene_indicios_representacion_judicial": record[
                    "tiene_indicios_representacion_judicial"
                ],
            }
        )
        for item in record["nombramientos"]:
            appointment_rows.append({"archivo": Path(record["archivo"]).name, **item})

    excel_path = output_dir / "camara_extraction.xlsx"
    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Resumen", index=False)
            pd.DataFrame(appointment_rows).to_excel(writer, sheet_name="Nombramientos", index=False)
    except PermissionError as exc:
        raise SystemExit(
            f"No pude escribir el Excel porque parece estar abierto o bloqueado: {excel_path}. "
            "Cierra el archivo y vuelve a correr el script."
        ) from exc


def find_pdfs(inputs: list[str]) -> list[Path]:
    pdfs: list[Path] = []
    for raw in inputs:
        path = Path(raw)
        if path.is_dir():
            pdfs.extend(sorted(path.glob("*.pdf")))
        elif path.suffix.lower() == ".pdf":
            pdfs.append(path)
    return pdfs


# ===========================================================================
# SECCIÓN 2 — Utilidades DOCX compartidas (verbatim de pre_fill.ipynb)
# ===========================================================================
def fill_placeholders(xml_bytes: bytes, replacements: dict[str, str]) -> bytes:
    """
    Reemplaza placeholders del tipo [razon_social] y [nit_cam] en el XML.

    Primera pasada: reemplazo directo sobre el texto XML.
      → Funciona cuando el placeholder está íntegro en un solo <w:t>.

    Segunda pasada: opera SOLO dentro de la región separate→end de cada campo
      FORMTEXT, sin tocar los runs del label ni la estructura del campo.
      → Cubre el caso en que Word partió el placeholder en varios runs.
    """
    # ── Pasada 1: reemplazo directo ───────────────────────────────────────────
    xml_str = xml_bytes.decode("utf-8")
    for placeholder, value in replacements.items():
        xml_str = xml_str.replace(placeholder, value)
    xml_bytes = xml_str.encode("utf-8")

    # ── Pasada 2: runs partidos dentro de campos FORMTEXT ────────────────────
    root = etree.fromstring(xml_bytes)

    for sep_run in root.xpath(
        "//w:r[w:fldChar[@w:fldCharType='separate']]",
        namespaces={"w": W_NS},
    ):
        parent = sep_run.getparent()
        siblings = list(parent)
        sep_idx = siblings.index(sep_run)

        # Recorre siblings siguientes hasta fldChar end
        value_runs: list[tuple] = []
        for elem in siblings[sep_idx + 1 :]:
            if elem.tag != f"{{{W_NS}}}r":
                continue
            if elem.xpath("w:fldChar[@w:fldCharType='end']", namespaces={"w": W_NS}):
                break
            t = elem.find(f"{{{W_NS}}}t")
            if t is not None:
                value_runs.append((elem, t))

        if not value_runs:
            continue

        full_text = "".join(t.text or "" for _, t in value_runs)
        changed = False
        for placeholder, value in replacements.items():
            if placeholder in full_text:
                full_text = full_text.replace(placeholder, value)
                changed = True

        if changed:
            _, first_t = value_runs[0]
            first_t.text = full_text
            first_t.set(f"{{{XML_NS}}}space", "preserve")
            for _, t in value_runs[1:]:
                t.text = ""

    return etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)


def create_filled_docx(
    template_path: Path,
    razon_social: str,
    nit: str,
    output_path: Path,
) -> None:
    """Copia la plantilla y reemplaza [razon_social] y [nit]."""
    replacements = {
        "[razon_social]": razon_social,
        "[nit_cam]": nit,
    }
    with zipfile.ZipFile(template_path, "r") as zin:
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "word/document.xml":
                    data = fill_placeholders(data, replacements)
                zout.writestr(item, data)


def safe_filename(name: str) -> str:
    """Elimina caracteres no permitidos en nombres de archivo."""
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


# --- CRS FATCA Form ---
def build_crs_fields(row: dict) -> tuple[dict, dict]:
    """Devuelve (campos_pag4, campos_pag11) a partir de una fila del Excel."""
    domicilio = str(row.get("domicilio_principal", "")).strip()
    is_bogota = domicilio == BOGOTA

    p4 = {
        "Text16": str(row.get("razon_social", "")).strip(),
        "Text17": "Colombia" if is_bogota else "",
        "Text18": str(row.get("direccion_domicilio_principal", "")).strip(),
        "Text19": "Bogota"   if is_bogota else "",
        "Text20": "Bogota"   if is_bogota else "",
        "Text21": "Colombia" if is_bogota else "",
        "Text22": "111111"   if is_bogota else "",
    }
    p11 = {
        "CJTaxResidence1": "Colombia" if is_bogota else "",
        "TIN1":            str(row.get("nit", "")).strip(),
    }
    return p4, p11


def fill_crs_pdf(template_path: Path, p4: dict, p11: dict, output_path: Path) -> None:
    """Rellena los campos del CRS FATCA Form y guarda en output_path."""
    reader = PdfReader(str(template_path))
    writer = PdfWriter()
    writer.append(reader)
    writer.update_page_form_field_values(writer.pages[3],  p4)
    writer.update_page_form_field_values(writer.pages[10], p11)
    with open(output_path, "wb") as f:
        writer.write(f)


# --- Formulario Grandes Exposiciones ---
def parse_nit(nit_raw: str) -> tuple[str, str]:
    """
    Separa el NIT en (número, dígito_verificación).
    Ejemplo: '860034594-1' → ('860034594', '1')
    Si no hay guion devuelve (nit_raw, '').
    """
    if "-" in nit_raw:
        parts = nit_raw.split("-", 1)
        return parts[0].strip(), parts[1].strip()
    return nit_raw.strip(), ""


def fill_grandes_exposiciones(template_path: Path, row: dict, output_path: Path) -> None:
    """Copia el template y rellena las celdas C5–C11."""
    shutil.copy2(str(template_path), str(output_path))
    wb = load_workbook(str(output_path))
    ws = wb.active

    razon_social = str(row.get("razon_social", "")).strip()
    nit_raw      = str(row.get("nit", "")).strip()
    nit_num, nit_dig = parse_nit(nit_raw)

    # C5 — Fecha de diligenciamiento (DD/MM/YYYY)
    ws["C5"] = date.today()
    ws["C5"].number_format = "DD/MM/YYYY"

    # C6 — Nombre / Razón Social
    ws["C6"] = razon_social

    # C7 — Tipo ID (dropdown: NIT | TIN | OTHER)
    ws["C7"] = "NIT" if nit_raw else None

    # C8 — Dígito de verificación (dropdown: 1-9 | not applicable.)
    if nit_dig:
        try:
            ws["C8"] = int(nit_dig)
        except ValueError:
            ws["C8"] = nit_dig

    # C9 — Número de identificación (dígitos antes del guion)
    if nit_num:
        try:
            ws["C9"] = int(nit_num)
        except ValueError:
            ws["C9"] = nit_num

    # C10 — Nombre Representante Legal (placeholder)
    ws["C10"] = "[por favor diligenciar]"

    # C11 — Identificación Representante Legal (placeholder)
    ws["C11"] = "[por favor diligenciar]"

    wb.save(str(output_path))


# --- GCC Declaración de Veracidad ---
def fill_gcc_pdf(template_path: Path, razon_social: str, nit: str, output_path: Path) -> None:
    """Rellena los campos AcroForm del GCC-Declaracion de Veracidad."""
    reader = PdfReader(str(template_path))
    writer = PdfWriter()
    writer.append(reader)
    writer.update_page_form_field_values(writer.pages[0], {
        "Nombre de la sociedad/empresa": razon_social,
        "(NIT)":                         nit,
    })
    with open(output_path, "wb") as f:
        writer.write(f)


# --- KYC application form ---
def build_kyc_fields(row: dict) -> dict:
    """
    Construye el diccionario de campos AcroForm a partir de una fila del Excel.

    Mapeo de campos:
      Fecha Date               ← Hoy  (DD/MM/YYYY)
      Text2  [nit_cam]         ← nit             (col C)
      Text3  [razon_social]    ← razon_social     (col B)
      Text4  [Teléfono…]       ← Teléfono comercial 1: (col F)
      Text5  [address]         ← direccion_domicilio_principal (col E)
      Text6  [pais]            ← "Colombia" si domicilio == "Bogotá D.C.", si no ""
      Código CIIU ISIC Code    ← Actividad principal Código CIIU (col G)
    """
    domicilio    = str(row.get("domicilio_principal", "")).strip()
    is_bogota    = domicilio == BOGOTA
    telefono_raw = row.get("Teléfono comercial 1:", "")
    ciiu_raw     = row.get("Actividad principal Código CIIU", "")

    return {
        "Fecha Date":            date.today().strftime("%d/%m/%Y"),
        "Text2":                 str(row.get("nit", "")).strip(),
        "Text3":                 str(row.get("razon_social", "")).strip(),
        "Text4":                 str(telefono_raw).strip() if pd.notna(telefono_raw) else "",
        "Text5":                 str(row.get("direccion_domicilio_principal", "")).strip(),
        "Text6":                 "Colombia" if is_bogota else "",
        "Código CIIU ISIC Code": str(ciiu_raw).strip() if pd.notna(ciiu_raw) else "",
    }


def fill_kyc_pdf(template_path: Path, fields: dict, output_path: Path) -> None:
    """Rellena los campos AcroForm del KYC form en todas las páginas del PDF."""
    reader = PdfReader(str(template_path))
    writer = PdfWriter()
    writer.append(reader)
    for page in writer.pages:
        writer.update_page_form_field_values(page, fields)
    with open(output_path, "wb") as f:
        writer.write(f)


# ===========================================================================
# ORQUESTADORES DE ALTO NIVEL (nuevos, usados por la GUI / el .exe)
# ===========================================================================
def run_extraction(pdf_paths, output_dir, log: Callable[[str], None] = print) -> Path:
    """Extrae datos de los PDF y escribe camara_extraction.xlsx/.json en output_dir."""
    output_dir = Path(output_dir)
    pdfs = find_pdfs([str(p) for p in pdf_paths])
    if not pdfs:
        raise FileNotFoundError("No se encontraron PDFs para procesar.")
    log(f"Procesando {len(pdfs)} PDF(s)...")
    records = []
    for pdf in pdfs:
        rec = extract_pdf(pdf)
        records.append(rec)
        log(f"  {Path(rec['archivo']).name}: {rec['razon_social']} | "
            f"NIT {rec['nit']} | {len(rec['nombramientos'])} nombramientos")
    write_outputs(records, output_dir)
    excel_path = output_dir / "camara_extraction.xlsx"
    log(f"Excel generado: {excel_path}")
    return excel_path


def run_prefill(excel_path, templates_dir, output_dir,
                log: Callable[[str], None] = print) -> int:
    """Lee el Excel y genera los 7 documentos por empresa en output_dir."""
    excel_path = Path(excel_path)
    templates_dir = Path(templates_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(excel_path, sheet_name="Resumen")
    log(f"{len(df)} empresa(s) cargada(s) desde {excel_path.name}")
    tpl = {k: templates_dir / v for k, v in TEMPLATE_FILES.items()}

    total = 0
    for _, row in df.iterrows():
        razon_social = str(row.get("razon_social", "")).strip()
        nit = str(row.get("nit", "")).strip()
        if not razon_social:
            log("  [SKIP] Fila sin razón social")
            continue
        base = safe_filename(razon_social)
        log(f"\n-> {razon_social}  (NIT {nit})")

        create_filled_docx(tpl["annex"], razon_social, nit,
                           output_dir / f"{base} - Annex A.docx")
        log("   OK  Annex A")

        create_filled_docx(tpl["contact"], razon_social, nit,
                           output_dir / f"{base} - Contact Information Certificate.docx")
        log("   OK  Contact Information Certificate")

        p4, p11 = build_crs_fields(row)
        fill_crs_pdf(tpl["crs"], p4, p11,
                     output_dir / f"{base} - CRS FATCA Form.pdf")
        log("   OK  CRS FATCA Form")

        create_filled_docx(tpl["esst"], razon_social, nit,
                           output_dir / f"{base} - Electronic Signatures Services Terms.docx")
        log("   OK  Electronic Signatures Services Terms")

        fill_grandes_exposiciones(tpl["ge"], row,
                                  output_dir / f"{base} - Formulario Grandes Exposiciones.xlsx")
        log("   OK  Formulario Grandes Exposiciones")

        fill_gcc_pdf(tpl["gcc"], razon_social, nit,
                     output_dir / f"{base} - GCC-Declaracion de Veracidad.pdf")
        log("   OK  GCC-Declaracion de Veracidad")

        fields = build_kyc_fields(row)
        fill_kyc_pdf(tpl["kyc"], fields,
                     output_dir / f"{base} - KYC application form.pdf")
        log("   OK  KYC application form")
        total += 1

    log(f"\n{total} empresa(s) procesada(s). Documentos en: {output_dir}")
    return total


def create_counterparty_folders(excel_path, network_base, output_dir,
                                log: Callable[[str], None] = print) -> None:
    """Crea en la red la carpeta de cada contraparte y copia los documentos a 'Only MO'."""
    excel_path = Path(excel_path)
    network_base = Path(network_base)
    output_dir = Path(output_dir)

    df = pd.read_excel(excel_path, sheet_name="Resumen")
    for _, row in df.iterrows():
        razon_social = str(row.get("razon_social", "")).strip()
        nit = str(row.get("nit", "")).strip()
        if not razon_social:
            log("  [SKIP] Fila sin razón social")
            continue
        folder_name = _safe_folder_name(f"{razon_social} - {nit}")
        counterparty_dir = network_base / folder_name
        if counterparty_dir.exists():
            raise FileExistsError(
                f"El folder ya existe y no se sobreescribirá: {counterparty_dir}"
            )
        only_mo = counterparty_dir / "Only MO"
        only_maker = counterparty_dir / "Only Maker"
        only_mo.mkdir(parents=True)
        only_maker.mkdir(parents=True)
        n = 0
        for file in output_dir.iterdir():
            if file.is_file():
                shutil.copy2(file, only_mo / file.name)
                n += 1
        log(f"  Folder creado: {folder_name}   (Only MO: {n} archivo(s))")
